import openpyxl
label_standard_path='E:\\KGlearning\\20200103\\Sample\\标准症状_修改_新(1)(1).xlsx'
sum_data_path='E:\\KGlearning\\20200103\\Sample\\griddle-jqcheck.xlsx'
chaifen_path='chaifensum.xlsx'
firstfile_path='E:\\KGlearning\\20200103\\Sample\\stand_zz_pair.xlsx'


def label_standard():
    labelexcel=openpyxl.load_workbook(label_standard_path)
    labelws=labelexcel['标准对应表']
    labeldict={}
    for i in range(2,labelws.max_row+1):
        label=labelws.cell(i,2).value
        if label not in labeldict and label!= None:
            labeldict[label]=[]
    print(len(labeldict))
    for i in range(2,labelws.max_row+1):
        label = labelws.cell(i, 2).value
        if label != None:
            standard=labelws.cell(i,4).value
            if standard not in labeldict[label]:
                labeldict[label].append(standard)
    for key,value in labeldict.items():
        print(key,value)

    return labeldict

def dividesum():

    labeldict=label_standard()
    sumexcel=openpyxl.load_workbook(sum_data_path)
    sumws=sumexcel['sum']
    sumws1=sumexcel.create_sheet('sum(标准症状不在同标签下)')
    sumws2 = sumexcel.create_sheet('sum(标准症状在同标签下)')

    recordrow1=2
    recordrow2=2
    print(sumws.max_row)
    for i in range(2,sumws.max_row+1):
        if sumws.cell(i,2).value!='' and sumws.cell(i,2).value!=None:
            #获取所有标准症状词
            standardlist=[]
            for j in range(3,sumws.max_column+1):
                if sumws.cell(i,j).value!=''and sumws.cell(i,j).value!=None:
                    standardlist.append(sumws.cell(i,j).value)

            #设置标志位
            biaozhi=0
            if len(standardlist)==1:
                biaozhi=0
            else:
                for key,value in labeldict.items():
                    if len(list(set(value).intersection(set(standardlist))))>=2:
                        biaozhi=1

            #标志为0则表示标准症状词都在不同标签下
            if biaozhi==0:
                print(standardlist)
                sumws1.cell(row=recordrow1, column=1, value=sumws.cell(i, 1).value)
                sumws1.cell(row=recordrow1, column=2, value=sumws.cell(i, 2).value)
                for n in range(len(standardlist)):
                    sumws1.cell(row=recordrow1, column=3+n, value=standardlist[n])
                recordrow1+=1
            else:
                print(standardlist)
                sumws2.cell(row=recordrow2, column=1, value=sumws.cell(i, 1).value)
                sumws2.cell(row=recordrow2, column=2, value=sumws.cell(i, 2).value)
                for n in range(len(standardlist)):
                    sumws2.cell(row=recordrow2, column=3 + n, value=standardlist[n])
                recordrow2+=1

    sumexcel.save(sum_data_path)

def chaifen():
    sumexcel = openpyxl.load_workbook(sum_data_path)
    sumws = sumexcel['sum(标准症状不在同标签下)']
    chaifenexcel=openpyxl.Workbook()
    sumws1_ = chaifenexcel.create_sheet('sum(（拆解）标准症状不在同标签下)')

    recordrow=1
    for i in range(2,sumws.max_row+1):
        standardlist = []
        if sumws.cell(i,2).value!='' and sumws.cell(i,2).value!=None:
            #获取所有标准症状词
            for j in range(3,sumws.max_column+1):
                if sumws.cell(i,j).value!=''and sumws.cell(i,j).value!=None:
                    standardlist.append(sumws.cell(i,j).value)
        for n in range(len(standardlist)):
            sumws1_.cell(row=recordrow, column=1, value=sumws.cell(i, 2).value)
            sumws1_.cell(row=recordrow, column=2, value=standardlist[n])
            recordrow+=1
    chaifenexcel.save(chaifen_path)

def shaixuan():
    firstfile=openpyxl.load_workbook(firstfile_path)
    fws=firstfile['pair']
    sumws1 = firstfile.create_sheet('标准症状不在同标签下')
    sumws2 = firstfile.create_sheet('标准症状在同标签下')
    pairdict={}
    for i in range(1,fws.max_row+1):
        print(i)
        yuanshi=fws.cell(i,1).value
        biaozhun=fws.cell(i,2).value
        if yuanshi not in pairdict:
            pairdict[yuanshi]=[]
            pairdict[yuanshi].append(biaozhun)
        else:
            pairdict[yuanshi].append(biaozhun)

    labeldict=label_standard()
    mrow=1
    nrow=1
    for yuanshi,biaozhun in pairdict.items():
        biaozhun=list(set(biaozhun))
        biaozhi=0
        if len(biaozhun)==1:
            biaozhi=0
        else:
            for key, value in labeldict.items():
                if len(list(set(value).intersection(set(biaozhun)))) >= 2:
                    biaozhi = 1

        if biaozhi==0:
            sumws1.cell(row=mrow, column=1, value=yuanshi)
            for n in range(len(biaozhun)):
                sumws1.cell(row=mrow, column=2+n, value=biaozhun[n])
            mrow+=1
        else:
            sumws2.cell(row=nrow, column=1, value=yuanshi)
            for n in range(len(biaozhun)):
                sumws2.cell(row=nrow, column=2 + n, value=biaozhun[n])
            nrow+=1

    firstfile.save(firstfile_path)

shaixuan()
# list1=['两耳蝉鸣', '耳如蝉鸣', '两耳轰鸣']
# list2=['两耳蝉鸣', '耳如']
# result=list(set(list1).intersection(set(list2)))
# print(result)
