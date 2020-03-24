# -*- coding:utf-8 -*-
import Levenshtein
import numpy as np
import openpyxl


embedding_path='E:\\桌面\\中医\\搬砖\\zhongyi_zi300.vector'
standard_path='E:\\KGlearning\\20200103\\label_standard.xlsx'
data_path='E:\\KGlearning\\20200103\\stand_zz_pair.xlsx'
resultpath='C:\\Users\\taohu\\Desktop\\备份\\sample_part6.xlsx'
alllabel=['情绪', '神志', '运化', '呼吸', '泌尿生殖', '胁', '头', '腹', '胸', '腰', '筋', '脉', '肉', '皮', '骨', '目', '乳房', '舌', '口', '肛门', '鼻', '咽喉', '耳', '生殖器']


vector_dic={}
vacab=[]
file = open(embedding_path,'r',encoding='utf-8')
embed = file.readlines()
for i in range(1,len(embed)):
    word1 = embed[i].replace('\n','').split(' ')[0]
    vec1 = embed[i].replace('\n','').split(' ')[1:]
    for i in range(len(vec1)):
        vec1[i]=float(vec1[i])
    vec1 = np.array(vec1)
    vacab.append(word1)
    vector_dic[word1]=vec1

def cosinesim(word1,word2):
    vec1 = np.zeros(300)
    vec2 = np.zeros(300)
    for str1 in word1:
        if str1 in vacab:
            vec1 += vector_dic[str1]
    for str2 in word2:
        if str2 in vacab:
            vec2 += vector_dic[str2]

    sim3 = float(np.dot(vec1, vec2) / (np.linalg.norm(vec1) * np.linalg.norm(vec2)))
    return sim3


def findlabelstandard(word):
    standardexcel=openpyxl.load_workbook(standard_path)
    standardwb=standardexcel['1']
    all=[]
    for row in standardwb.iter_rows(min_row=1):
        row_data = []
        for cell in row:
            if cell.value!=None and cell.value not in alllabel:
                row_data.append(str(cell.value))
        if word in row_data:
            all=row_data
            break
    return all

page=1
dataexcel=openpyxl.load_workbook(data_path)
datawb=dataexcel['pair']
resultexcel=openpyxl.Workbook()
resultwb = resultexcel.create_sheet('sim'+str(page))

resultwb.cell(row=1, column=1, value='原始症状X')
resultwb.cell(row=1, column=2, value='标准症状Y')
resultwb.cell(row=1, column=3, value='匹配得分（Jaro（X）*0.25+余弦相似度（X）*0.25+Jaro（Y）*0.25+余弦相似度（Y）*0.25）*0.9')
resultwb.cell(row=1, column=4, value='JaroX')
resultwb.cell(row=1, column=5, value='余弦相似度X')
resultwb.cell(row=1, column=6, value='JaroY')
resultwb.cell(row=1, column=7, value='余弦相似度Y')

j=2
#for k in range(1,datawb.max_row+1):
for k in range(13001,datawb.max_row+1):
    standword=str(datawb.cell(k,2).value)
    yuanshi=str(datawb.cell(k,1).value)

    all=findlabelstandard(standword)
    #print(all)
    print(k)
    #print(len(all))
    if j == 1040001:
        page += 1
        resultwb = resultexcel.create_sheet('sim' + str(page))
        j = 2
    resultwb.cell(row=j, column=1, value=yuanshi)
    resultwb.cell(row=j, column=2, value=standword)
    resultwb.cell(row=j, column=3, value=1)
    j+=1
    for i in range(len(all)):
        if standword!=str(all[i]):
            jarox = Levenshtein.jaro_winkler(yuanshi, str(all[i]))
            jaroy = Levenshtein.jaro_winkler(standword, str(all[i]))
            cosix=cosinesim(yuanshi, str(all[i]))*0.5+0.5
            cosiy = cosinesim(standword, str(all[i]))*0.5+0.5
            if j==1040001:
                page+=1
                resultwb = resultexcel.create_sheet('sim'+str(page))
                j=2
            resultwb.cell(row=j, column=1, value=yuanshi)
            resultwb.cell(row=j, column=2, value=str(all[i]))
            #middleresult=str(jarox)+'|'+str(cosix)+'|'+str(jaroy)+'|'+str(cosiy)
            #resultwb.cell(row=j, column=i * 3 + 5, value=middleresult)
            resultwb.cell(row=j, column=4, value=jarox)
            resultwb.cell(row=j, column=5, value=cosix)
            resultwb.cell(row=j, column=6, value=jaroy)
            resultwb.cell(row=j, column=7, value=cosiy)
            score=(jarox*0.25+jaroy*0.25+cosix*0.25+cosiy*0.25)*0.9
            resultwb.cell(row=j, column=3, value=score)
            j+=1


resultexcel.save(resultpath)




